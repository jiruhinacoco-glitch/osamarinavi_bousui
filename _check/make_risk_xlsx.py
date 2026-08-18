# -*- coding: utf-8 -*-
# 納まりナビ：セキュリティリスク・課題の整理表（初心者向け・日本語）
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = 'Yu Gothic'
GREEN = '1C6B3C'; GREEN_L = 'E8F3E6'; CREAM = 'FFFDF7'
RED = 'C0392B'; RED_L = 'FADBD8'
ORG = 'B9770E'; ORG_L = 'FDEBD0'
BLU = '1A5276'; BLU_L = 'D6EAF8'
GRN2 = '1E8449'; GRN2_L = 'D5F5E3'

thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def base(c, size=10.5, bold=False, color='222222'):
    c.font = Font(name=F, size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    c.border = BORDER

def head(c):
    c.font = Font(name=F, size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=GREEN)
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    c.border = BORDER

def level(c, text):
    c.value = text
    m = {'高': (RED, RED_L), '中': (ORG, ORG_L), '低': (BLU, BLU_L), '良い点': (GRN2, GRN2_L)}
    col, fill = m.get(text.split('（')[0], ('222222', 'FFFFFF'))
    c.font = Font(name=F, size=10.5, bold=True, color=col)
    c.fill = PatternFill('solid', fgColor=fill)
    c.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    c.border = BORDER

wb = Workbook()

# ============ シート1：まとめ ============
ws = wb.active
ws.title = 'まとめ（先に読む）'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 92

r = 2
c = ws.cell(r, 2, '納まりナビ　セキュリティの現状整理（2026-08-06時点）')
c.font = Font(name=F, size=16, bold=True, color=GREEN)
r += 1
c = ws.cell(r, 2, 'このファイルは佐野さんに直接お渡しするもので、公開の場（GitHub）には置いていません。')
c.font = Font(name=F, size=10, color='888888')
r += 2

c = ws.cell(r, 2, '先に結論（3行）')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=GREEN)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
points = [
    ('①', 'お客さんが入力したデータ（工程表・図面・写真など）は端末の中だけに保存され、外に送っていません。この点は安全です。'),
    ('②', 'いちばんの心配は「見られること」。アプリの画面・プログラム・開発メモが、いまは世界中の誰でも読める状態です（検索には出ない設定にはしてある）。'),
    ('③', '特に、現場記録帳のサンプル77件（住所・発注元・数量・単価・金額入り）が実際の台帳由来なら、単価や粗利の水準が競合や元請に見えてしまいます。ここが最優先の確認事項です。'),
]
for mark, tx in points:
    c = ws.cell(r, 2, mark)
    c.font = Font(name=F, size=11, bold=True, color=GREEN)
    c.alignment = Alignment(vertical='top', horizontal='center')
    c.border = BORDER
    c2 = ws.cell(r, 3, tx)
    base(c2, 11)
    c2.fill = PatternFill('solid', fgColor=CREAM)
    ws.row_dimensions[r].height = 34
    r += 1
r += 1

c = ws.cell(r, 2, 'まず知っておく言葉（これだけ）')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=GREEN)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
terms = [
    ('公開リポジトリ', 'プログラム一式の置き場（GitHub）。いまは「公開」設定なので、アプリの中身も、作り方を書いた開発メモも、誰でも読めます。'),
    ('GitHub Pages', 'いまサイトを載せている無料の公開サービス。URLを知っている人は誰でも開けます（ログインの仕組みはありません）。'),
    ('端末内の保存（localStorage）', 'ブラウザ（Safari等）がスマホの中に持っている保存箱。入力データはここに入ります。ネットには送られない反面、ブラウザのデータ消去で一緒に消えます。'),
    ('APIキー', '地図（Googleマップ）を使うための利用券のような文字列。他人に渡ると、自分の名義で地図を使われてしまう恐れがあります。'),
    ('検索避け（noindex）', '「検索結果に出さないでください」という印。GoogleやBingには出なくなりますが、URLを直接知っている人には効きません。'),
    ('変更履歴（コミット履歴）', 'プログラムの「いつ何を直したか」の記録。一度公開したものは、消した後も履歴に残り、さかのぼって見ることができます。'),
]
for name, desc in terms:
    c = ws.cell(r, 2, name)
    base(c, 10.5, bold=True, color=GREEN)
    c.fill = PatternFill('solid', fgColor=GREEN_L)
    c2 = ws.cell(r, 3, desc)
    base(c2)
    ws.row_dimensions[r].height = 30
    r += 1
r += 1

c = ws.cell(r, 2, '「危険度」の見かた')
c.font = Font(name=F, size=12, bold=True, color='FFFFFF')
c.fill = PatternFill('solid', fgColor=GREEN)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
r += 1
legend = [
    ('高', '商売に直接ひびく恐れがある。早めに手を打ちたい。', RED, RED_L),
    ('中', 'すぐ困るわけではないが、放っておくといつか痛い。', ORG, ORG_L),
    ('低', '念のため知っておけばよい。', BLU, BLU_L),
    ('良い点', 'いまの作りで既に守れているところ。', GRN2, GRN2_L),
]
for name, desc, col, fill in legend:
    c = ws.cell(r, 2, name)
    c.font = Font(name=F, size=10.5, bold=True, color=col)
    c.fill = PatternFill('solid', fgColor=fill)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDER
    c2 = ws.cell(r, 3, desc)
    base(c2)
    r += 1

# ============ シート2：リスク一覧 ============
ws2 = wb.create_sheet('リスク一覧')
ws2.sheet_view.showGridLines = False
headers = ['No', '分野', 'いまどうなっているか', '何が心配か（起きうること）', '危険度',
           'いまできている対策', 'これからの対策（提案）', '費用・手間の目安']
widths = [4, 15, 38, 38, 8, 30, 38, 16]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    head(ws2.cell(1, i, h))
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
ws2.row_dimensions[1].height = 24

rows = [
    ('サンプルデータ',
     '現場記録帳に、住所・発注元・工法・数量・単価・金額の入った77件のサンプルが入っています。手元のExcel台帳をもとに作られたものです。あわせて、労務単価25,000円/人日などの数字もプログラム内に書かれています。',
     'もし実在の現場・実際の単価に近い値なら、単価と粗利の水準が競合や元請に丸見えになります。見積交渉で不利になる恐れがあります。',
     '高',
     'なし（サンプルとして表示されている状態）',
     '【最優先で確認】77件が実データ由来かどうかを確認する。実データに近いなら、住所・金額・単価を架空の値に差し替える（プログラムの構造は変えずにデータだけ差し替えられます）。',
     '確認は今すぐ・0円\n差し替えは半日'),
    ('見られる範囲',
     'サイトはURLを知っている人なら誰でも開けます。ログイン（合言葉で入る仕組み）はありません。検索避けは設定済みなので、GoogleやBingの検索には出ません。',
     'URLが人づてに広まると、競合他社が画面や機能をそのまま見て真似できます。',
     '高',
     '検索避け（noindex）を全ページに設定済み。SNS等で宣伝しない方針も決定済み。',
     'Cloudflare Pages＋Access（無料）へ引っ越し、「許可した人（50人まで無料）だけが開ける」形にする。※引っ越しでデータが消える問題は、今回作った「データの保存・復元」で対応済み。',
     '0円・作業1日\n＋許可する人の登録'),
    ('プログラムの公開',
     'プログラムの置き場（リポジトリ）が「公開」設定です。アプリの中身に加えて、開発メモ（なぜそう作ったか・今後の計画・営業方針まで書いたファイル）も誰でも読めます。',
     '機能を真似されるより、「なぜそうしたか」「これから何を作るか」を読まれるほうが痛い、という状態です。',
     '高',
     '開発メモから個人情報（自宅PCの情報など）は削除済み。',
     '案A：GitHub Pro（月4ドル≒600円）でリポジトリを非公開にする（ただしサイト自体は見られるまま）。\n案B：Cloudflareへ引っ越した後にリポジトリを非公開にする（0円・こちらが本命）。',
     '案A：月600円・10分\n案B：上の引っ越しと同時'),
    ('過去の履歴',
     '一度消した個人的なメモ（端末の設定の取り決めの詳細・自宅PCのフォルダ名）が、変更履歴をさかのぼると今も読めます。',
     '個人的な事情・環境の情報が第三者に知られる可能性があります。電話番号やメールアドレスは含まれていないことを確認済みです。',
     '中',
     '最新版からは削除済み（履歴にだけ残っている）。',
     '履歴の書き換え（過去の記録から該当部分を完全に消す作業）。10分程度で終わりますが、履歴が作り直されるため、佐野さんの了解をもらってから実施します。',
     '0円・10分\n※了解待ち'),
    ('地図のAPIキー',
     '現場マップで使うGoogleマップの利用券（APIキー）を、画面から入力して端末内に保存する作りです。キーそのものはプログラムには書かれていません（良い作り）。',
     'もしキーが人に渡り、そのキーに「使える場所の制限」がかかっていないと、他人が自分の名義で地図を使い、Googleからの請求が増える恐れがあります。',
     '中',
     'キーをプログラムに書き込まず、各端末で入力する方式にしてある。',
     'Googleの管理画面で、キーに「このサイトのURLからしか使えない」制限（リファラ制限）をかける。これで漏れても他所では使えなくなります。',
     '0円・15分'),
    ('バックアップファイル',
     '今回作った「データの保存・復元」で書き出すファイルには、全データ（写真・APIキー含む）が1つにまとまって入ります。',
     'このファイルをLINEやメールで人に送ると、中身がまるごと相手に渡ります。',
     '中',
     '読み込み時は「納まりナビのファイルかどうか」を確認し、知らないデータは書き込まない作りにしてある。',
     'ファイルは自分の保管場所（iPhoneの「ファイル」やパソコン）だけに置き、人に送らない。この一覧表と一緒に運用ルールとして覚えておく。',
     '0円・習慣のみ'),
    ('データが消える',
     '入力データは端末の中だけにあります。外に漏れない反面、Safariの「履歴とWebサイトデータを消去」・端末の容量不足・スマホの紛失や故障で、全部消えます。',
     'せっかく入力した工程表・図面・写真が一瞬で消え、取り戻せない。（漏れる話ではなく「失う」話ですが、実害はこれが一番起きやすい）',
     '中',
     '「データの保存・復元」機能を作成済み（2026-08-06d・ホームの「設定」から）。',
     '月に1回など、区切りのタイミングでファイルに書き出す習慣をつける。将来的には自動でサーバーに保存する仕組み（要サーバー・有料）も検討。',
     '0円・月1回の操作'),
    ('アカウント乗っ取り',
     'サイトの元となるGitHubのアカウントが乗っ取られると、サイトの中身を書き換えられたり消されたりします。',
     'サイト改ざん・削除。パスワードの使い回しがあると危険度が上がります。',
     '中',
     '（GitHub側の設定状況は未確認）',
     '2段階認証（ログイン時にスマホでの確認を追加する設定）をオンにする。パスワードは他サイトと使い回さない。',
     '0円・10分'),
    ('外部の部品',
     '3D表示の部品（three.js）だけ、表示のたびに外部の配布サイトから読み込んでいます。それ以外はすべて自分のサイト内で完結しています。',
     '配布サイト側に問題が起きた場合に3D表示が動かなくなる・（理論上は）細工された部品を読まされる可能性。実際に起きる確率は低いです。',
     '低',
     '読み込めなかったときは案内を出して止まる作りにしてある。',
     '部品のファイル（約600KB）を自分のサイト内に置いて、外部から読まない形にする。オフラインでも3Dが動くようになるおまけ付き。',
     '0円・30分'),
    ('通信の暗号化',
     'サイトは https（通信の暗号化）で配信されています。途中で盗み見・改ざんされない仕組みです。',
     '—（現状で問題ありません）',
     '良い点',
     'GitHub Pagesが自動で対応。引っ越し先候補のCloudflareでも同様に自動対応。',
     'このままでOK。',
     '—'),
    ('入力データの扱い',
     'お客さんが入力したデータを外部に送信する処理は一切ありません。集計も図面もPDFも、全部端末の中だけで作っています。',
     '—（現状で問題ありません。「漏れる場所がない」のが強みです）',
     '良い点',
     'アプリの設計そのものが「端末内で完結」になっている。',
     'このままでOK。将来サーバー保存を入れるときに改めて設計する。',
     '—'),
]
for i, (field, now, worry, lv, done, todo, cost) in enumerate(rows, start=1):
    rr = i + 1
    c = ws2.cell(rr, 1, i); base(c, 10.5, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='top')
    c = ws2.cell(rr, 2, field); base(c, 10.5, bold=True)
    base(ws2.cell(rr, 3, now))
    base(ws2.cell(rr, 4, worry))
    level(ws2.cell(rr, 5), lv)
    base(ws2.cell(rr, 6, done))
    base(ws2.cell(rr, 7, todo))
    base(ws2.cell(rr, 8, cost))
    if lv == '良い点':
        for col in (1, 2, 3, 4, 6, 7, 8):
            ws2.cell(rr, col).fill = PatternFill('solid', fgColor='F4FBF4')
    ws2.row_dimensions[rr].height = 96

# ============ シート3：やることの順番 ============
ws3 = wb.create_sheet('やることの順番')
ws3.sheet_view.showGridLines = False
headers3 = ['順番', 'やること', '費用', '手間', 'ひとこと']
widths3 = [6, 52, 14, 14, 52]
for i, (h, w) in enumerate(zip(headers3, widths3), start=1):
    head(ws3.cell(1, i, h))
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'
ws3.row_dimensions[1].height = 22

plan = [
    ('サンプル77件が「実データ由来かどうか」を教えてください', '0円', '返事だけ',
     'これが一番大事な確認です。実データに近いなら、住所・単価・金額を架空の値に差し替えます（半日でできます）。'),
    ('GitHubの2段階認証をオンにする', '0円', '10分',
     'スマホでの確認を1つ足すだけで、乗っ取りにかなり強くなります。設定のお手伝いができます。'),
    ('地図のAPIキーに「このサイトからしか使えない」制限をかける', '0円', '15分',
     'Googleの管理画面での設定です。手順を案内できます。'),
    ('過去の履歴から個人的なメモを完全に消す（了解をもらえれば実施）', '0円', '10分',
     '前々からの「判断待ち」の1つ目です。OKの返事をもらえたらすぐやります。'),
    ('Cloudflare Pages＋Accessへ引っ越し、「許可した人だけ見られるサイト」にする', '0円', '1日',
     '判断待ちの2つ目・本命です。引っ越しでデータが消える問題は「データの保存・復元」で解決済みなので、いつでも動けます。'),
    ('引っ越し完了後、GitHubのリポジトリを非公開にする', '0円', '10分',
     'これでプログラムも開発メモも外から見えなくなり、「見られる」系のリスクがほぼ解消します。'),
    ('月1回、ホームの「設定」からデータを書き出して保管する（習慣）', '0円', '月5分',
     'スマホの故障・紛失に備える保険です。書き出したファイルは人に送らないこと。'),
]
for i, (act, cost, effort, note) in enumerate(plan, start=1):
    rr = i + 1
    c = ws3.cell(rr, 1, i); base(c, 11, bold=True, color=GREEN)
    c.alignment = Alignment(horizontal='center', vertical='top')
    base(ws3.cell(rr, 2, act), )
    ws3.cell(rr, 2).font = Font(name=F, size=10.5, bold=True)
    c = ws3.cell(rr, 3, cost); base(c); c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    c = ws3.cell(rr, 4, effort); base(c); c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    base(ws3.cell(rr, 5, note))
    ws3.row_dimensions[rr].height = 42

out = '/tmp/claude-0/-home-user-osamarinavi-bousui/14021813-84a1-5c50-96c6-9a3eca49d2e3/scratchpad/納まりナビ_セキュリティ整理_2026-08-06.xlsx'
wb.save(out)
print('saved:', out)
